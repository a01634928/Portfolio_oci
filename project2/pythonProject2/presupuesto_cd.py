import openpyxl
from openpyxl import load_workbook

"""
def active(self):  # property
    return self._sheets[self._active_sheet_index]


def active(self, value):  # active.setter
    self._active_sheet_index = value"""


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    porcentaje = 1.312
    wb = load_workbook(filename='PPTO_CD_CONSTITUCION_2021.xlsx')
    wb.active = 0
    ws = wb.active
    #Servicios preventivos
    for i in range(10, 29):
        for j in range(6, 9, 2):
            cel = ws.cell(row=i, column=j)
            cel_value = cel.internal_value
            cel_change = float(cel_value[2:6])
            cel_change = round(cel_change * porcentaje, 2)

            cel_value = cel_value[:2] + str(cel_change) + cel_value[6:]
            print(cel_value)
            cel.value = cel_value
    #Enero-Diciembre
    for i in range(10, 29):  # rows
        for j in range(9, 33):  # columns
            cel = ws.cell(row=i, column=j)
            cel_value = cel.internal_value
            if cel_value is None:
                print('non')
            else:
                cel_change = float(cel_value)
                cel_value = round(cel_change * porcentaje, 2)
                cel.value = cel_value
                print(cel_value)
# Worksheet 2-8
    for k in range(1, 9):
        wb.active = k
        ws = wb.active
        # Enero-Diciembre
        for i in range(10, 29):  # rows
            for j in range(7, 31):  # columns
                cel = ws.cell(row=i, column=j)
                cel_value = cel.internal_value
                if cel_value is None or cel_value == '':
                    print('non')
                else:
                    cel_change = float(cel_value)
                    cel_value = round(cel_change * porcentaje, 2)
                    cel.value = cel_value
                    print(cel_value)
        wb.save("PPTO_CD_CONSTITUCION_2021_1.xlsx")


