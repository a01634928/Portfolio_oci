import openpyxl
from openpyxl import load_workbook
#14-95 Worsheet 11
"""
def active(self):  # property
    return self._sheets[self._active_sheet_index]


def active(self, value):  # active.setter
    self._active_sheet_index = value"""


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    porcentaje = 0.825
    wb = load_workbook(filename='PPTO _LORETO_2021_1.xlsx')
    wb.active = 0
    ws = wb.active
    #Servicios preventivos
    for i in range(10, 15):
        for j in range(6, 9, 2):
            cel = ws.cell(row=i, column=j)
            cel_value = cel.internal_value
            if cel_value is None or 0 or '0':
                print('non')
            else:
                cel_change = float(cel_value[2:6])
                cel_change = round(cel_change * porcentaje, 2)
                cel_value = cel_value[:2] + str(cel_change) + cel_value[6:]
                print(cel_value)
                cel.value = cel_value
    #Enero-Diciembre
    for i in range(10, 15):  # rows
        for j in range(9, 33):  # columns
            cel = ws.cell(row=i, column=j)
            cel_value = cel.internal_value
            if cel_value is None:
                print('non')
            else:
                cel_change = float(cel_value)
                cel_value = round(cel_change * porcentaje, 2)
                cel.value = cel_value
                print(cel_value)
# Worksheet 2-8
    for k in range(1, 9):
        wb.active = k
        ws = wb.active
        # Enero-Diciembre
        for i in range(10, 15):  # rows
            for j in range(7, 31):  # columns
                cel = ws.cell(row=i, column=j)
                cel_value = cel.internal_value
                if cel_value is None:
                    print('non')
                else:
                    cel_change = float(cel_value)
                    cel_value = round(cel_change * porcentaje, 2)
                    cel.value = cel_value
                    print(cel_value)
    """                    
        wb.active = 10
        ws = wb.active
        #cel = ws['O85']
        #cel = ws.cell(row=85, column=15, values_only=True)
        #cel = cel.value  # Valor esperado 113,000
        cel = ws["O85"].value

        print('internal_value: ' + str(cel))
        if cel <= 113000:
            contador = 1
        print('internal_value: ' + str(cel))"""
    wb.save("PPTO _LORETO_2021_2.xlsx")